# Title: This script include a class to design a review database page.
#
# Author: Farhad Abdollahi (farhad.abdollahi.ctr@dot.gov)
# Date: 11/25/2024
# ======================================================================================================================

# Importing the required libraries.
import os
import sys
import sqlite3
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from matplotlib import cm
from matplotlib.colors import to_hex, LinearSegmentedColormap
from PyQt5.QtWidgets import (QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QGroupBox,
                             QHBoxLayout, QVBoxLayout, QPushButton, QWidget, QMessageBox, QLabel, QFormLayout, 
                             QComboBox, QPlainTextEdit, QInputDialog, QFileDialog)
from PyQt5.QtGui import QFont, QBrush, QColor
from PyQt5.QtCore import Qt
from scripts.Sub02_CreateNewSQLTable import Get_DB_SummaryData, Get_Identifier_Combinations
from scripts.Sub04_FTIR_Analysis_Functions import Binary_to_Array

# Define the custom cmap for the table COV colors.
Reds = cm.get_cmap('Reds', 256)             # Get the "reds" colormap.
# Clip at 0.85 to prevent excessive dark red colors.
New = Reds(np.linspace(0, 0.75, 256))
Custom_cmap = LinearSegmentedColormap.from_list("custom_reds", New)


class DB_ReviewPage(QMainWindow):
    """
    This class design a review database page in which the database is shown as table where user can select any specific 
    sample and revise its analysis procedure. 
    """

    def __init__(self, conn, cursor, DB_Name, DB_Folder, stack, shared_data):
        # Initiate the required parameters.
        super().__init__()
        self.conn = conn            # connection to the SQL database.
        self.cursor = cursor        # cursor for running the SQL commands.
        self.DB_Name = DB_Name      # Name of the database.
        self.DB_Folder = DB_Folder  # Directory at which the database is saved.
        self.stack = stack
        self.shared_data = shared_data
        self.ColumnNames = ['id', 'ID-number', 'Lab Aging', 'Rep #',
                            'ICO (Decon)', 'ISO (Decon)',
                            'ICO (base)', 'ICO (tang)', 'ISO (base)', 'ISO (tang)',
                            'Carbonyl Peak loc (1/cm)', 'Sulfoxide Peak loc (1/cm)',
                            'Aliphatic Peak loc 1 (1/cm)', 'Aliphatic Peak loc 2 (1/cm)',
                            'Carbonyl Peak val', 'Sulfoxide Peak val',
                            'Aliphatic Peak val 1', 'Aliphatic Peak val 2',
                            'Carbonyl Area (base)', 'Sulfoxide Area (base)', 'Aliphatic Area (base)',
                            'Carbonyl Area (tang)', 'Sulfoxide Area (tang)', 'Aliphatic Area (tang)',
                            'Carbonyl Xmin', 'Carbonyl Xmax',
                            'Sulfoxide Xmin', 'Sulfoxide Xmax',
                            'Aliphatic Xmin', 'Aliphatic Xmax']
        self.SQL_ColumnNames = [
            'id', 'Bnumber', 'Lab_Aging', 'RepNumber',
            'Deconv_ICO', 'Deconv_ISO',
            'ICO_Baseline', 'ICO_Tangential', 'ISO_Baseline', 'ISO_Tangential',
            'Carbonyl_Peak_Wavenumber', 'Sulfoxide_Peak_Wavenumber',
            'Aliphatic_Peak_Wavenumber_1', 'Aliphatic_Peak_Wavenumber_2',
            'Carbonyl_Peak_Absorption', 'Sulfoxide_Peak_Absorption',
            'Aliphatic_Peak_Absorption_1', 'Aliphatic_Peak_Absorption_2',
            'Carbonyl_Area_Baseline',   'Sulfoxide_Area_Baseline',   'Aliphatic_Area_Baseline',
            'Carbonyl_Area_Tangential', 'Sulfoxide_Area_Tangential', 'Aliphatic_Area_Tangential',
            'Carbonyl_Min_Wavenumber', 'Carbonyl_Max_Wavenumber',
            'Sulfoxide_Min_Wavenumber', 'Sulfoxide_Max_Wavenumber',
            'Aliphatic_Min_Wavenumber', 'Aliphatic_Max_Wavenumber']
        self.ColumnNamesAnalysis = [
            'B_Number', 'Lab_Aging_Condition', 'Num_Data',
            'ICO_Baseline_mean', 'ICO_Baseline_std', 'ICO_Baseline_COV',
            'ICO_Deconv_mean', 'ICO_Deconv_std', 'ICO_Deconv_COV',
            'ICO_Tangential_mean', 'ICO_Tangential_std', 'ICO_Tangential_COV',
            'ISO_Baseline_mean', 'ISO_Baseline_std', 'ISO_Baseline_COV',
            'ISO_Deconv_mean', 'ISO_Deconv_std', 'ISO_Deconv_COV',
            'ISO_Tangential_mean', 'ISO_Tangential_std', 'ISO_Tangential_COV',
            'Aliphatic_Area_Baseline_mean', 'Aliphatic_Area_Baseline_std', 'Aliphatic_Area_Baseline_COV',
            'Aliphatic_Area_Tangential_mean', 'Aliphatic_Area_Tangential_std', 'Aliphatic_Area_Tangential_COV',
            'Carbonyl_Peak_Wavenumber_mean', 'Carbonyl_Peak_Wavenumber_std', 'Carbonyl_Peak_Wavenumber_COV',
            'Sulfoxide_Peak_Wavenumber_mean', 'Sulfoxide_Peak_Wavenumber_std', 'Sulfoxide_Peak_Wavenumber_COV',
            'Carbonyl_Peak_Absorption_mean', 'Carbonyl_Peak_Absorption_std', 'Carbonyl_Peak_Absorption_COV',
            'Sulfoxide_Peak_Absorption_mean', 'Sulfoxide_Peak_Absorption_std', 'Sulfoxide_Peak_Absorption_COV']
        self.IdentifierCombs = Get_Identifier_Combinations(self.cursor)
        self.PushButtonStyle = {
            "General": """
        QPushButton:enabled {
            background-color: #DCECF9; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:disabled {
            background-color: #dcdcdc; color: #808080;border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:hover {
            background-color: lightgray; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:pressed {
            background-color: gray; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        """, 
            "Export" : """
        QPushButton:enabled {
            background-color: #93E9BE; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:disabled {
            background-color: #dcdcdc; color: #808080;border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:hover {
            background-color: #56C28B; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:pressed {
            background-color: gray; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:checked {
            background-color: lime; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        """, 
            "Delete":  """
        QPushButton:enabled {
            background-color: #FA8072; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:disabled {
            background-color: #dcdcdc; color: #808080;border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:hover {
            background-color: #CC5F55; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:pressed {
            background-color: gray; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:checked {
            background-color: lime; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        """, 
            "Modify": """
        QPushButton:enabled {
            background-color: #F2E394; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:disabled {
            background-color: #dcdcdc; color: #808080;border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:hover {
            background-color: #E0D57D; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:pressed {
            background-color: gray; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        QPushButton:checked {
            background-color: lime; color: #000; border: 1px solid #B0B0B0; border-radius: 8px; padding: 6px 12px;}
        """, }
        self.initUI()
    # ------------------------------------------------------------------------------------------------------------------
    def initUI(self):
        # # Initiate the user interface.
        # self.setWindowTitle(f"AutoFTIR (version 1.0) | Database name: {self.DB_Name}")
        # self.setFixedSize(1500, 900)
        # Main widget and layout
        main_widget = QWidget()
        # Make a horizontal layout.
        layout = QHBoxLayout(main_widget)
        self.setCentralWidget(main_widget)
        # self.setStyleSheet("background-color: #f0f0f0;")
        # Generate the left and right layouts, where left one include summay info and plots, and right one include the
        #   buttons, adjustments, and controls.
        LeftLayout = QVBoxLayout()
        RightLayout = QVBoxLayout()
        # --------------------------------------------------------------------------------------------------------------
        # Section 1: summary information.
        SummaryData = Get_DB_SummaryData(self.cursor)
        Section01 = QGroupBox("Summary Information")
        Section01.setStyleSheet("QGroupBox { font-weight: bold; }")
        Section01_Layout = QHBoxLayout()
        FormLayout_Left  = QFormLayout()            # Define a form layout for the left side.
        FormLayout_Right = QFormLayout()            # Define a form layout for the right side. 
        # Create the left side labels in Section 01.
        Label01 = QLabel("Number of data:")
        self.Label_NumData = QLabel(f'{SummaryData["NumRows"]}')
        Label02 = QLabel("Number of valid data:")
        self.Label_NumValidData = QLabel(f'{SummaryData["NumValidRows"]}')
        Label03 = QLabel("Avg Number Replicates:")
        self.Label_AvgNumReplicates = QLabel(f'{SummaryData["AvgNumRep"]:.1f}')
        # Create the right side labels in Section 01.
        Label04 = QLabel("Number of unique B-numbers:")
        self.Label_NumUniqueBnum = QLabel(f'{SummaryData["NumUniqueBnumber"]}')
        Label05 = QLabel("Number of unique Lab agings:")
        self.Label_NumUniqueLabAge = QLabel(f'{SummaryData["NumUniqueLabAging"]}')
        Label06 = QLabel("Number of Unique Bnum/LabAge:")
        self.Label_NumUniqueBnumLabAge = QLabel(f'{SummaryData["NumUniqueBnumLabAge"]}')
        # Syncing button.
        self.Button_Sync = QPushButton("Refresh summay")
        self.Button_Sync.setFont(QFont("Arial", 8))
        self.Button_Sync.clicked.connect(self.Sync_Summary_Info)
        self.Button_Sync.setStyleSheet(self.PushButtonStyle['General'])
        self.Button_Sync.setSizePolicy(self.Button_Sync.sizePolicy().Expanding, self.Button_Sync.sizePolicy().Preferred)
        # Place the labels in the GUI.
        FormLayout_Left.addRow(Label01,  self.Label_NumData)
        FormLayout_Left.addRow(Label02,  self.Label_NumValidData)
        FormLayout_Left.addRow(Label03,  self.Label_AvgNumReplicates)
        FormLayout_Right.addRow(Label04, self.Label_NumUniqueBnum)
        FormLayout_Right.addRow(Label05, self.Label_NumUniqueLabAge)
        FormLayout_Right.addRow(Label06, self.Label_NumUniqueBnumLabAge)
        Section01_Layout.addLayout(FormLayout_Left)
        Section01_Layout.addLayout(FormLayout_Right)
        Section01_Layout.addWidget(self.Button_Sync)
        Section01.setLayout(Section01_Layout)
        LeftLayout.addWidget(Section01, 10)
        # --------------------------------------------------------------------------------------------------------------
        # Section 02: Table to show the records.
        Section02 = QGroupBox("Show Database")
        Section02.setStyleSheet("QGroupBox { font-weight: bold; }")
        Section02_Layout = QVBoxLayout()
        # Adding one label to show the number of fetched data.
        self.Label_NumFetchedRows = QLabel('Number of fetched data (rows): 0')
        # Create the table
        self.Table = QTableWidget()
        self.Table.setRowCount(10)
        self.Table.setColumnCount(len(self.ColumnNames))
        self.Table.setHorizontalHeaderLabels(self.ColumnNames)
        self.Table.setSelectionBehavior(self.Table.SelectRows)
        self.Table.setSelectionMode(self.Table.SingleSelection)
        # Placing the table in the window.
        Section02_Layout.addWidget(self.Label_NumFetchedRows)
        Section02_Layout.addWidget(self.Table)
        Section02.setLayout(Section02_Layout)
        LeftLayout.addWidget(Section02, 90)
        # --------------------------------------------------------------------------------------------------------------
        # Section 3: filtering the database results.
        Section03 = QGroupBox("Filtering Options")
        Section03.setStyleSheet("QGroupBox { font-weight: bold; }")
        Section03_Layout = QVBoxLayout()
        # First, dropdown for the B-numbers.
        Label_DropDown01 = QLabel("Select the asphalt binder B-number:")
        self.DropDown_Bnumber = QComboBox()
        self.DropDown_Bnumber.addItems(['Please Select...', 'All binders'] +
                                       list(self.IdentifierCombs['Bnumber'].unique()))
        self.DropDown_Bnumber.currentIndexChanged.connect(
            self.Function_DropDown_Bnumber)
        # Second dropdown for the Laboratory aging state.
        Label_DropDown02 = QLabel("Select the laboratory aging level:")
        self.DropDown_LabAging = QComboBox()
        self.DropDown_LabAging.addItems(['All Aging Levels'])
        self.DropDown_LabAging.setEnabled(False)
        self.DropDown_LabAging.currentIndexChanged.connect(
            self.Function_DropDown_LabAging)
        # Finally, add the apply button.
        self.Button_Fetch = QPushButton("Fetch data")
        self.Button_Fetch.setStyleSheet(self.PushButtonStyle['General'])
        self.Button_Fetch.setSizePolicy(self.Button_Fetch.sizePolicy().Expanding, 
                                        self.Button_Fetch.sizePolicy().Preferred)
        self.Button_Fetch.clicked.connect(self.Function_Button_Fetch)
        # Place the items in the window.
        Section03_Layout.addWidget(Label_DropDown01)
        Section03_Layout.addWidget(self.DropDown_Bnumber)
        Section03_Layout.addWidget(Label_DropDown02)
        Section03_Layout.addWidget(self.DropDown_LabAging)
        Section03_Layout.addWidget(self.Button_Fetch)
        Section03.setLayout(Section03_Layout)
        RightLayout.addWidget(Section03, 20)
        # --------------------------------------------------------------------------------------------------------------
        # Section 04:
        Section04 = QGroupBox("Operation")
        Section04.setStyleSheet("QGroupBox { font-weight: bold; }")
        Section04_Layout = QVBoxLayout()
        # First button for the modification of the current selection.
        self.Button_Modify = QPushButton("Modify Selected Record")
        self.Button_Modify.setStyleSheet(self.PushButtonStyle['Modify'])
        self.Button_Modify.clicked.connect(self.Function_Button_Modify)
        self.Button_Modify.setSizePolicy(self.Button_Modify.sizePolicy().Expanding, 
                                         self.Button_Modify.sizePolicy().Preferred)
        # Second button to change the view to Analysis of DB results.
        self.Button_Analysis = QPushButton("Analysis Results Page")
        self.Button_Analysis.setStyleSheet(self.PushButtonStyle['General'])
        self.Button_Analysis.clicked.connect(self.Function_Button_Analysis)
        self.Button_Analysis.setSizePolicy(self.Button_Analysis.sizePolicy().Expanding, 
                                           self.Button_Analysis.sizePolicy().Preferred)
        # Next button for going back to the previous page.
        self.Button_Go2Main = QPushButton("Main Page")
        self.Button_Go2Main.setStyleSheet(self.PushButtonStyle['General'])
        self.Button_Go2Main.clicked.connect(self.Function_Button_Go2Main)
        self.Button_Go2Main.setSizePolicy(self.Button_Go2Main.sizePolicy().Expanding, 
                                          self.Button_Go2Main.sizePolicy().Preferred)
        # Next button for Exporting an individual record.
        self.Button_Export_Record = QPushButton("Export Individual Record")
        self.Button_Export_Record.setStyleSheet(self.PushButtonStyle['Export'])
        self.Button_Export_Record.clicked.connect(self.Function_Button_Export_Individual)
        self.Button_Export_Record.setSizePolicy(self.Button_Export_Record.sizePolicy().Expanding, 
                                                self.Button_Export_Record.sizePolicy().Preferred)
        # Next button for Exporting the whole database.
        self.Button_Export_Database = QPushButton("Export Database Summary")
        self.Button_Export_Database.setStyleSheet(self.PushButtonStyle['Export'])
        self.Button_Export_Database.clicked.connect(self.Function_Button_Export_Database)
        self.Button_Export_Database.setSizePolicy(self.Button_Export_Database.sizePolicy().Expanding, 
                                                  self.Button_Export_Database.sizePolicy().Preferred)
        # Next button for Deleting a record.
        self.Button_Delete_Record = QPushButton("Delete Selected Record")
        self.Button_Delete_Record.setStyleSheet(self.PushButtonStyle['Delete'])
        self.Button_Delete_Record.clicked.connect(self.Function_Button_Delete_Record)
        self.Button_Delete_Record.setSizePolicy(self.Button_Delete_Record.sizePolicy().Expanding, 
                                                self.Button_Delete_Record.sizePolicy().Preferred)
        # Placement of the buttons.
        Section04_Layout.addWidget(self.Button_Modify)
        Section04_Layout.addWidget(self.Button_Delete_Record)
        Section04_Layout.addWidget(self.Button_Export_Record)
        Section04_Layout.addWidget(self.Button_Export_Database)
        Section04_Layout.addWidget(self.Button_Analysis)
        Section04_Layout.addWidget(self.Button_Go2Main)
        Section04.setLayout(Section04_Layout)
        RightLayout.addWidget(Section04, 30)
        # --------------------------------------------------------------------------------------------------------------
        # Section 04: Output section.
        Section05 = QGroupBox("Output section (Terminal-like)")
        Section05.setStyleSheet("QGroupBox { font-weight: bold; }")
        Section05_Layout = QVBoxLayout()
        self.Terminal = QPlainTextEdit(self)
        # Make it read-only. User shouldn't write!
        self.Terminal.setReadOnly(True)
        self.Terminal.setStyleSheet("background-color: black; color: white;")
        self.Terminal.appendPlainText(">>> Review_Database_Results()\n")
        Section05_Layout.addWidget(self.Terminal)
        Section05.setLayout(Section05_Layout)
        RightLayout.addWidget(Section05, 50)
        # --------------------------------------------------------------------------------------------------------------
        # Final placement of the right and left layouts.
        layout.addLayout(LeftLayout, 70)
        layout.addLayout(RightLayout, 30)
    # ------------------------------------------------------------------------------------------------------------------
    def ShowEvent(self, event):
        # Update the identifier combinations.
        self.IdentifierCombs = Get_Identifier_Combinations(self.cursor)
        self.Function_Button_Fetch()
    # ------------------------------------------------------------------------------------------------------------------
    # Define the functions.
    def Sync_Summary_Info(self):
        # Get the latest summary information.
        SummaryData = Get_DB_SummaryData(self.cursor)
        # Update the values.
        self.Label_NumData.setText(f'{SummaryData["NumRows"]}')
        self.Label_NumValidData.setText(f'{SummaryData["NumValidRows"]}')
        self.Label_AvgNumReplicates.setText(f'{SummaryData["AvgNumRep"]:.1f}')
        self.Label_NumUniqueBnum.setText(f'{SummaryData["NumUniqueBnumber"]}')
        self.Label_NumUniqueLabAge.setText(
            f'{SummaryData["NumUniqueLabAging"]}')
        self.Label_NumUniqueBnumLabAge.setText(
            f'{SummaryData["NumUniqueBnumLabAge"]}')
        # Return Nothing.
        return
    # ------------------------------------------------------------------------------------------------------------------
    def Function_DropDown_Bnumber(self):
        # First, empty the table.
        self.Table.clearSelection()
        self.Label_NumFetchedRows.setText('Number of fetched data (rows): 0')
        self.Table.setRowCount(10)
        for row_idx in range(10):
            for col_idx in range(len(self.ColumnNames)):
                self.Table.setItem(row_idx, col_idx, QTableWidgetItem(''))
        # Determine the state of the LabAging dropdown menu.
        if self.DropDown_Bnumber.currentIndex() == 0:
            # "Please select..." is selected.
            self.DropDown_LabAging.setEnabled(False)
            self.DropDown_LabAging.setCurrentIndex(0)
            self.DropDown_LabAging.clear()
            self.DropDown_LabAging.addItems(['All Aging Levels'])
        elif self.DropDown_Bnumber.currentIndex() == 1:
            # "All binders" is selected.
            self.DropDown_LabAging.clear()
            self.DropDown_LabAging.addItems(
                ['All Aging Levels'] + list(self.IdentifierCombs['Lab_Aging'].unique()))
            self.DropDown_LabAging.setCurrentIndex(0)
            self.DropDown_LabAging.setEnabled(True)
        else:
            # A specific B-number was selected.
            Bnumber = self.DropDown_Bnumber.currentText()
            TempDF = self.IdentifierCombs[self.IdentifierCombs['Bnumber'] == Bnumber]
            self.DropDown_LabAging.clear()
            self.DropDown_LabAging.addItems(
                ['All Aging Levels'] + list(TempDF['Lab_Aging'].unique()))
            self.DropDown_LabAging.setCurrentIndex(0)
            self.DropDown_LabAging.setEnabled(True)
        # Return nothing.
        return
    # ------------------------------------------------------------------------------------------------------------------
    def Function_DropDown_LabAging(self):
        # Empty the table.
        self.Table.clearSelection()
        self.Label_NumFetchedRows.setText('Number of fetched data (rows): 0')
        self.Table.setRowCount(10)
        for row_idx in range(10):
            for col_idx in range(len(self.ColumnNames)):
                self.Table.setItem(row_idx, col_idx, QTableWidgetItem(''))
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Button_Fetch(self):
        # First, read the filters.
        self.Table.clearSelection()
        Bnumber = self.DropDown_Bnumber.currentText()
        LabAging = self.DropDown_LabAging.currentText()
        if self.DropDown_Bnumber.currentIndex() == 0:
            return
        elif self.DropDown_Bnumber.currentIndex() == 1:         # For all binders.
            if self.DropDown_LabAging.currentIndex() == 0:    # For all aging levels.
                self.cursor.execute(
                    f"SELECT {', '.join(self.SQL_ColumnNames)} FROM FTIR")
            else:
                self.cursor.execute(f"SELECT {', '.join(self.SQL_ColumnNames)} FROM FTIR WHERE Lab_Aging = ?",
                                    (LabAging,))
        else:       # For a specific binder.
            if self.DropDown_LabAging.currentIndex() == 0:    # for all aging levels.
                self.cursor.execute(f"SELECT {', '.join(self.SQL_ColumnNames)} FROM FTIR WHERE Bnumber = ?",
                                    (Bnumber,))
            else:
                self.cursor.execute(f"SELECT {', '.join(self.SQL_ColumnNames)} FROM FTIR " +
                                    f"WHERE Bnumber = ? AND Lab_Aging = ?", (Bnumber, LabAging))
        # Get the rows.
        Rows = self.cursor.fetchall()
        # Modify the table with the row value
        self.Label_NumFetchedRows.setText(
            f'Number of fetched data (rows): {len(Rows)}')
        self.Table.setRowCount(len(Rows))
        for row_idx, row_data in enumerate(Rows):
            for col_idx, cell_data in enumerate(row_data):
                if type(cell_data) == float:
                    item = QTableWidgetItem(f'{cell_data:.4f}')
                else:
                    item = QTableWidgetItem(str(cell_data))
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.Table.setItem(row_idx, col_idx, item)
        # Print the selection to the terminal.
        Msg = f'>>> "{len(Rows)}" rows found in the Database, for:\n>>>\tBnumber: {Bnumber}\n>>>\tAging: {LabAging}'
        self.Terminal.appendPlainText(Msg)
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Button_Go2Main(self):
        """
        This function simply clear the current page and move to the main page.
        """
        # Clear the table and drop down menus.
        self.Table.clearSelection()
        self.DropDown_Bnumber.setCurrentIndex(0)
        self.DropDown_LabAging.setCurrentIndex(0)
        self.Table.setRowCount(10)
        for row_idx in range(10):
            for col_idx in range(len(self.ColumnNames)):
                self.Table.setItem(row_idx, col_idx, QTableWidgetItem(''))
        # Move to the main page.
        self.stack.setCurrentIndex(0)
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Button_Modify(self):
        # Find the selected index. 
        idx, ID = self.Check_Row_Selection(ActionLabel='modify')
        if (ID == -1) or (idx == -1):
            return
        # Call Stack 3 and send the ID.
        self.shared_data.data = ID
        self.stack.setCurrentIndex(2)
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Button_Analysis(self):
        # First check which view needed to be shown.
        if self.Button_Analysis.text() == "Show Analysis Results":
            # First of all, check if the analysis is available.
            self.cursor.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name = ?", ("FTIR_Analysis_DB",))
            Response = self.cursor.fetchone()
            if not Response:
                # Such table is not existed and user need to first analyze the results.
                QMessageBox.critical(self, "Analysis NOT available",
                                     f"The analysis results are NOT available. Please use the 'Main Page' button and " +
                                     f"go back to the main page. Then, click on 'Analyze DB and Export to Excel' " +
                                     f"button to perform the analysis and then use this page for visualization.")
                return
            # If results are available, prepare the page.
            self.DropDown_Bnumber.setEnabled(False)
            self.DropDown_Bnumber.setCurrentIndex(0)
            self.DropDown_LabAging.setEnabled(False)
            self.DropDown_LabAging.setCurrentIndex(0)
            self.Button_Fetch.setEnabled(False)
            self.Button_Modify.setEnabled(False)
            self.Button_Go2Main.setEnabled(False)
            self.Terminal.appendPlainText(
                f"\n>>> Moving to the Analysis of Results view:")
            # Get the results from the DB.
            self.cursor.execute(
                f"SELECT {', '.join(self.ColumnNamesAnalysis)} FROM FTIR_Analysis_DB")
            Rows = self.cursor.fetchall()
            # Clear the table.
            self.Table.clearContents()
            self.Table.clearSelection()
            self.Table.setColumnCount(len(self.ColumnNamesAnalysis))
            self.Table.setHorizontalHeaderLabels(self.ColumnNamesAnalysis)
            self.Table.setRowCount(len(Rows))
            # Fill the table with the new analysis results.
            for row_idx, row_data in enumerate(Rows):
                for col_idx, cell_data in enumerate(row_data):
                    if col_idx in [5, 8, 11, 14, 17, 20, 23, 26, 29, 32, 35, 38]:
                        try:
                            item = QTableWidgetItem(f'{cell_data*100:.1f}%')
                            ColorNumber = Get_Color_4_COV(cell_data)
                            item.setBackground(QBrush(QColor(ColorNumber)))
                        except:
                            item = QTableWidgetItem('None')
                    elif type(cell_data) == float:
                        item = QTableWidgetItem(f'{cell_data:.4f}')
                    else:
                        item = QTableWidgetItem(str(cell_data))
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                    self.Table.setItem(row_idx, col_idx, item)
            # Change the text on the button.
            self.Button_Analysis.setText('Show Database')
        else:
            # Show the database content.
            self.DropDown_Bnumber.setEnabled(True)
            self.DropDown_Bnumber.setCurrentIndex(0)
            self.Button_Modify.setEnabled(True)
            self.Button_Go2Main.setEnabled(True)
            self.Button_Fetch.setEnabled(True)
            self.Terminal.appendPlainText(f"\n>>> Moving to the DB view:")
            # Clear the table.
            self.Table.clearContents()
            self.Table.clearSelection()
            self.Table.setColumnCount(len(self.ColumnNames))
            self.Table.setHorizontalHeaderLabels(self.ColumnNames)
            self.Table.setRowCount(10)
            # change the name.
            self.Button_Analysis.setText('Show Analysis Results')
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Button_Delete_Record(self): 
        """
        This function deletes the current selection from the database. 
        """
        # Find the selected index. 
        idx, ID = self.Check_Row_Selection(ActionLabel='delete')
        if (ID == -1) or (idx == -1):
            return
        # SelectedIndices = self.Table.selectionModel().selectedIndexes()
        # if len(SelectedIndices) == 0:           
        #     # Nothing is selected. 
        #     QMessageBox.critical(self, "Data Selection Error!", 
        #                          f"Row was not selected. Please first select the row you want to delete " + 
        #                          f"from the database.")
        #     return
        # idx = SelectedIndices[0].row()
        # # Check the id value. 
        # ID = self.Table.item(idx, 0)
        # if ID == None or ID.text() == '':
        #     # Table is empty. 
        #     QMessageBox.critical(self, "Data Selection Error!",
        #                          f"Selected row ({idx + 1}) is empty. Please first fetch the data using the " +
        #                          f'"Search and Filter" section, then select the intended row, and then click ' +
        #                          f'"Delete Record" button.')
        #     return
        # else:
        #     ID = int(ID.text())
        Bnumber = self.Table.item(idx, 1).text()
        LabAging= self.Table.item(idx, 2).text()
        RepNum  = self.Table.item(idx, 3).text()
        Msg  = f'Do you want to Permanently Delete the following record?:\n' + \
                f'ID-number={Bnumber} at age level of "{LabAging}", Rep {RepNum}'
        Question = QMessageBox()
        Question.setIcon(QMessageBox.Question)
        Question.setWindowTitle("Delete Record Confirmation")
        Question.setText(Msg)
        Question.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        Question.setDefaultButton(QMessageBox.No)
        Reply = Question.exec_()
        # Check the response. 
        if Reply == QMessageBox.Yes:
            # User Choose Yes.
            # deleting the record from the Database.
            self.cursor.execute("DELETE FROM FTIR WHERE id = ?", (ID,))
            self.conn.commit()
            # Updating the table. 
            self.Function_Button_Fetch()
            # Print out the results in Terminal. 
            Msg = f'>>> 1 record was successfully deleted from Database:\n>>>\tID-number={Bnumber}\n' + \
                    f'>>>\tLab aging level: "{LabAging}"\n>>>\tRepetition number: {RepNum}\n>>>'
            self.Terminal.appendPlainText(Msg)
        else:
            # User clicked No, and ignore the deletion. 
            print("Ignore")
            return
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Button_Export_Database(self):
        """
        This function exports the whole database in an Excel file. It is noted that this button exports the database of 
        all individual records, and didn't combine the repititions of a sample. 
        """
        # Ask for a directory to save the file and file name. 
        Directory = QFileDialog.getExistingDirectory(self, "Please select Saving Directory", "")
        # If a file is selected by the user, update the Input_SavePath.
        if not Directory:
            QMessageBox.critical(self, "Directory Selection Failed!", f"Directory was NOT selected. Please try again.")
            return
        print(f'Saving Directory: {Directory}')
        # Ask for the file name. 
        FileName, IsOkButtonPressed = QInputDialog.getText(
            self, "Output File Name", "Please enter the output file name (without .xlsx):", 
            text=f'{self.DB_Name}_Export')
        if IsOkButtonPressed:
            FileName = FileName + '.xlsx'
            print(f"Saving File Name: {FileName}")
        else:
            QMessageBox.critical(self, "Output File Name Failed!", 
                                 f"Output file name was NOT confirmed. Please try again.")
            return
        # --------------------------------------------------------------------------------------------------------------
        # Define the exporting columns and their corresponding labels.
        ColNames = [
            'id', 'Bnumber', 'Lab_Aging', 'RepNumber', 'FileName', 'FileDirectory', 'IsOutlier',
            'Baseline_Adjustment_Method', 'ALS_Lambda', 'ALS_Ratio', 'ALS_NumIter',
            'Normalization_Method', 'Normalization_Coeff',
            'Deconv_ICO', 'Deconv_ISO',
            'ICO_Baseline', 'ISO_Baseline',
            'Carbonyl_Area_Baseline', 'Sulfoxide_Area_Baseline', 'Aliphatic_Area_Baseline',
            'ICO_Tangential', 'ISO_Tangential', 
            'Carbonyl_Area_Tangential', 'Sulfoxide_Area_Tangential', 'Aliphatic_Area_Tangential',
            'Carbonyl_Peak_Wavenumber', 'Sulfoxide_Peak_Wavenumber',
            'Aliphatic_Peak_Wavenumber_1', 'Aliphatic_Peak_Wavenumber_2',
            'Carbonyl_Peak_Absorption', 'Sulfoxide_Peak_Absorption',
            'Aliphatic_Peak_Absorption_1', 'Aliphatic_Peak_Absorption_2',
            'Carbonyl_Min_Wavenumber', 'Carbonyl_Max_Wavenumber',
            'Sulfoxide_Min_Wavenumber', 'Sulfoxide_Max_Wavenumber',
            'Aliphatic_Min_Wavenumber', 'Aliphatic_Max_Wavenumber']
        Labels = [
            'DB id', 'ID-number', 'Laboratory Aging', 'Repetition Number', 'File Name', 'File Directory', 'Is Outlier?',
            'Baseline Adjustment Method', 'ALS λ Coeff', 'ALS ρ Coeff', 'ALS Niter Coeff',
            'Normalization Method', 'Normalization Coeff',
            'ICO (deconvolution)', 'ISO (deconvolution)', 
            'ICO (baseline integration)', 'ISO (baseline integration)',
            'Carbonyl Area (baseline integration)', 'Sulfoxide Area (baseline integration)',
            'Aliphatic Area (baseline integration)',
            'ICO (tangential integration)', 'ISO (tangential integration)',
            'Carbonyl Area (tangential integration)', 'Sulfoxide Area (tangential integration)',
            'Aliphatic Area (tangential integration)',
            'Carbonyl Peak Wavenumber (cm⁻¹)', 'Sulfoxide Peak Wavenumber (cm⁻¹)',
            'Aliphatic Peak Wavenumber 1 (cm⁻¹)', 'Aliphatic Peak Wavenumber 2 (cm⁻¹)',
            'Carbonyl Peak Absorption', 'Sulfoxide Peak Absorption',
            'Aliphatic Peak Absorption 1', 'Aliphatic Peak Absorption 2',
            'Carbonyl Min Wavenumber', 'Carbonyl Max Wavenumber',
            'Sulfoxide Min Wavenumber', 'Sulfoxide Max Wavenumber', 
            'Aliphatic Min Wavenumber', 'Aliphatic Max Wavenumber']
        # --------------------------------------------------------------------------------------------------------------
        # Fetch data from Database. 
        self.cursor.execute(f"SELECT {','.join(ColNames)} FROM FTIR")
        Content = self.cursor.fetchall()
        # --------------------------------------------------------------------------------------------------------------
        # Prepare the output file. 
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        # Set the title of the sheet
        ws.title = "Sheet1"
        # Define some styles. 
        Title_fill      = PatternFill(start_color="FFE989", end_color="FFCC00", fill_type="solid")
        Valid_fill      = PatternFill(start_color="D4FEC2", end_color="D4FEC2", fill_type="solid")
        Invalid_fill    = PatternFill(start_color="FDBBBB", end_color="FDBBBB", fill_type="solid")
        Data_fill       = PatternFill(start_color="DFDED9", end_color="DFDED9", fill_type="solid")
        thin            = Side(border_style="thin", color="000000")
        cell_border     = Border(top=thin, left=thin, right=thin, bottom=thin)
        header_font     = Font(name="Arial", bold=True, size=11, color="000000")
        cell_font       = Font(name="Arial", size=11, color="000000")
        center_alignment= Alignment(horizontal="center", vertical="center")
        # Write the titles. 
        for j, col in enumerate(Labels, start=1):
            cell = ws.cell(row=1, column=j, value=col)
            cell.fill = Title_fill
            cell.border = cell_border
            cell.font = header_font
            cell.alignment = center_alignment
        # Write the data. 
        for i in range(len(Content)):
            Fill = Invalid_fill if Content[i][6] else Valid_fill
            for j, Value in enumerate(Content[i], start=1):
                cell = ws.cell(row=2 + i, column=j, value=Value)
                cell.fill = Fill
                cell.border = cell_border
                cell.font = cell_font
                cell.alignment = center_alignment
        # Adjust the size of each column in the final excel file. 
        for j, col in enumerate(Labels, start=1):
            col_letter = get_column_letter(j)
            # Estimate width: character count + padding (2 to 5 is typical)
            max_label_length = len(str(col)) + 2
            ws.column_dimensions[col_letter].width = max_label_length
        # Save the results. 
        wb.save(os.path.join(Directory, FileName))
        # Return nothing. 
        return
    # ------------------------------------------------------------------------------------------------------------------
    def Function_Button_Export_Individual(self):
        """
        This function exports the raw and analyzed results of the selected individual record as an Excel file. 
        """
        # Find the selected index. 
        idx, ID = self.Check_Row_Selection(ActionLabel='export')
        if (ID == -1) or (idx == -1):
            return
        # SelectedIndices = self.Table.selectionModel().selectedIndexes()
        # if len(SelectedIndices) == 0:           
        #     # Nothing is selected. 
        #     QMessageBox.critical(self, "Data Selection Error!", 
        #                                 f"Row was not selected. Please first select the row you want to export " + 
        #                                 f"from the database.")
        #     self.Export_ProgressBar.hide()
        #     return
        # idx = SelectedIndices[0].row()
        # # Check the id value. 
        # ID = self.Table.item(idx, 0)
        # if ID == None or ID.text() == '':
        #     # Table is empty. 
        #     QMessageBox.critical(self, "Data Selection Error!", 
        #                                 f"Selected row ({idx + 1}) is empty. Please first fetch the data using the " +
        #                                 f'"Search and Filter" section, then select the intended row, and then click ' +
        #                                 f'"Export (Individual Record)" button.')
        #     self.Export_ProgressBar.hide()
        #     return
        # # Otherwise, everything is ready for exporting. 
        # ID = int(ID.text())             # The ID to retrieve the data from the DB. 
        # --------------------------------------------------------------------------------------------------------------
        # Ask for a directory to save the file and file name. 
        Directory = QFileDialog.getExistingDirectory(self, "Please select Saving Directory", "")
        # If a file is selected by the user, update the Input_SavePath.
        if not Directory:
            QMessageBox.critical(self, "Directory Selection Failed!", f"Directory was NOT selected. Please try again.")
            return
        print(f'Saving Directory: {Directory}')
        # Ask for the file name. 
        self.cursor.execute("SELECT FileName FROM FTIR WHERE id = ?", (ID,))
        FileName, IsOkButtonPressed = QInputDialog.getText(
            self, "Output File Name", "Please enter the output file name (without .xlsx):", 
            text=os.path.splitext(self.cursor.fetchone()[0])[0])
        if IsOkButtonPressed:
            FileName = FileName + '.xlsx'
            print(f"Saving File Name: {FileName}")
        else:
            QMessageBox.critical(self, "Output File Name Failed!", 
                                 f"Output file name was NOT confirmed. Please try again.")
            return
        # --------------------------------------------------------------------------------------------------------------
        # Prepare the output file. 
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        # Set the title of the sheet
        ws.title = "Sheet1"
        # Define some styles. 
        Info_fill       = PatternFill(start_color="FFE989", end_color="FFCC00", fill_type="solid")
        Pre_fill        = PatternFill(start_color="A7E2FF", end_color="A7E2FF", fill_type="solid")
        Res_fill        = PatternFill(start_color="D4FEC2", end_color="D4FEC2", fill_type="solid")
        GL_fill         = PatternFill(start_color="FDBBBB", end_color="FDBBBB", fill_type="solid")
        Data_fill       = PatternFill(start_color="DFDED9", end_color="DFDED9", fill_type="solid")
        thin            = Side(border_style="thin", color="000000")
        cell_border     = Border(top=thin, left=thin, right=thin, bottom=thin)
        header_font     = Font(name="Arial", bold=True, size=11, color="000000")
        cell_font       = Font(name="Arial", size=11, color="000000")
        center_alignment= Alignment(horizontal="center", vertical="center")
        left_alignment  = Alignment(horizontal="left", vertical="center")
        # ----------------------------------------------------------------------------
        # Extract the General information parameters. 
        ColNames_Info = ['Bnumber', 'Lab_Aging', 'RepNumber', 'FileName',  'FileDirectory', 
                        'IsOutlier']
        Labels_Info   = ['B-number', 'Lab aging level', 'Repetition number', 'Raw data file name', 'Raw data file directory', 
                        'Is this test considered Outlier']
        self.cursor.execute(f'SELECT {", ".join(ColNames_Info)} FROM FTIR WHERE id = ?', (ID,))
        Values_Info = list(self.cursor.fetchone())
        # Make isOutlier yes/no.
        Values_Info[5] = 'Yes' if Values_Info[5] else "No"
        # ------------------------------------------------------
        # Write the General information. 
        ws.merge_cells('A1:B1')
        cell = ws.cell(row=1, column=1, value='General Information')
        cell.fill = Info_fill
        cell.border = cell_border
        cell.font = Font(name="Arial", size=13, bold=True, color="000000")
        cell.alignment = center_alignment
        for i, (header, value) in enumerate(zip(Labels_Info, Values_Info), start=2):
            # Write the title.
            cell1 = ws.cell(row=i, column=1, value=header + ':')
            cell1.fill = Info_fill
            cell1.border = cell_border
            cell1.font = header_font
            cell1.alignment = left_alignment
            # Write the value. 
            cell1 = ws.cell(row=i, column=2, value=value)
            cell1.fill = Info_fill
            cell1.border = cell_border
            cell1.font = cell_font
            cell1.alignment = left_alignment
        NextRowIndex = i + 2
        # ----------------------------------------------------------------------------------------------------------------------
        # Next, write the Pre-processing properties. 
        # Extract the results. 
        ColNames_Pre = ['Baseline_Adjustment_Method', 'ALS_Lambda', 'ALS_Ratio', 'ALS_NumIter', 
                        'Normalization_Method', 'Normalization_Coeff']
        Labels_Pre = ['Baseline adjustment method', 'ALSS λ coefficient', 'ALSS ρ coefficient', 'ALSS n coefficient', 
                    'Normalization method', 'Normalization β coefficient']
        self.cursor.execute(f'SELECT {", ".join(ColNames_Pre)} FROM FTIR WHERE id = ?', (ID,))
        Values_Pre = list(self.cursor.fetchone())
        # ------------------------------------------------------
        # Write the General information. 
        ws.merge_cells(f'A{NextRowIndex}:B{NextRowIndex}')
        cell = ws.cell(row=NextRowIndex, column=1, value='Pre-processing Properties')
        cell.fill = Pre_fill
        cell.border = cell_border
        cell.font = Font(name="Arial", size=13, bold=True, color="000000")
        cell.alignment = center_alignment
        for i in range(1, 9):
            for j in range(2):
                cell1 = ws.cell(row=NextRowIndex+i, column=j+1, value='')
                cell1.fill = Pre_fill
        # Put the image. 
        Image_Obj = Read_Resize_Image(ResourcePath(os.path.join(".", "assets", "ALSS Baseline Correction.png")), 153)
        ws.add_image(Image_Obj, f"A{NextRowIndex+1}")
        NextRowIndex += 9
        # Write the 2PP model parameters. 
        for i, (header, value) in enumerate(zip(Labels_Pre, Values_Pre), start=NextRowIndex):
            # Write the title.
            cell1 = ws.cell(row=i, column=1, value=header + ':')
            cell1.fill = Pre_fill
            cell1.border = cell_border
            cell1.font = header_font
            cell1.alignment = left_alignment
            # Write the value. 
            cell1 = ws.cell(row=i, column=2, value=value)
            cell1.fill = Pre_fill
            cell1.border = cell_border
            cell1.font = cell_font
            cell1.alignment = left_alignment
        NextRowIndex = i + 2
        # ----------------------------------------------------------------------------------------------------------------------
        # Next, provide the results. 
        # Extract the results. 
        ColNames_Res = [
            'Deconv_ICO', 'Deconv_ISO', 
            'ICO_Baseline', 'ISO_Baseline', 'ICO_Tangential', 'ISO_Tangential',
            'Carbonyl_Area_Baseline', 'Sulfoxide_Area_Baseline', 'Aliphatic_Area_Baseline', 
            'Carbonyl_Area_Tangential', 'Sulfoxide_Area_Tangential', 'Aliphatic_Area_Tangential',
            'Carbonyl_Peak_Wavenumber', 'Sulfoxide_Peak_Wavenumber', 
            'Aliphatic_Peak_Wavenumber_1', 'Aliphatic_Peak_Wavenumber_2',
            'Carbonyl_Peak_Absorption', 'Sulfoxide_Peak_Absorption', 
            'Aliphatic_Peak_Absorption_1', 'Aliphatic_Peak_Absorption_2',
            'Carbonyl_Min_Wavenumber', 'Carbonyl_Max_Wavenumber', 
            'Sulfoxide_Min_Wavenumber', 'Sulfoxide_Max_Wavenumber', 
            'Aliphatic_Min_Wavenumber', 'Aliphatic_Max_Wavenumber',]
        Labels_Res = [
            'ICO (deconvolution)', 'ISO (deconvolution)', 
            'ICO (baseline integration)', 'ISO (baseline integration)', 
            'ICO (tangential integration)', 'ISO (tangential integration)', 
            'Carbonyl area (baseline integration)', 
            'Sulfoxide area (baseline integration)', 
            'Aliphatic area (baseline integration)', 
            'Carbonyl area (tangential integration)', 
            'Sulfoxide area (tangential integration)', 
            'Aliphatic area (tangential integration)', 
            'Carbonyl peak location (cm⁻¹)', 'Sulfoxide peak location (cm⁻¹)', 
            'First Aliphatic peak location (cm⁻¹)', 'Second aliphatic peak location (cm⁻¹)',
            'Carbonyl peak absorbance', 'Sulfoxide peak absorbance', 
            'First Aliphatic peak absorbance', 'Second aliphatic peak absorbance', 
            'Carbonyl peak min boundary (cm⁻¹)', 'Carbonyl peak max boundary (cm⁻¹)', 
            'Sulfoxide peak min boundary (cm⁻¹)', 'Sulfoxide peak max boundary (cm⁻¹)', 
            'Aliphatic peak min boundary (cm⁻¹)', 'Aliphatic peak max boundary (cm⁻¹)',]
        self.cursor.execute(f'SELECT {", ".join(ColNames_Res)} FROM FTIR WHERE id = ?', (ID,))
        Values_Res = list(self.cursor.fetchone())
        # ------------------------------------------------------
        # Write the General information. 
        ws.merge_cells(f'A{NextRowIndex}:B{NextRowIndex}')
        cell = ws.cell(row=NextRowIndex, column=1, value='Results')
        cell.fill = Res_fill
        cell.border = cell_border
        cell.font = Font(name="Arial", size=13, bold=True, color="000000")
        cell.alignment = center_alignment
        NextRowIndex += 1
        # Write the Results. 
        for i, (header, value) in enumerate(zip(Labels_Res, Values_Res), start=NextRowIndex):
            # Write the title.
            cell1 = ws.cell(row=i, column=1, value=header + ':')
            cell1.fill = Res_fill
            cell1.border = cell_border
            cell1.font = header_font
            cell1.alignment = left_alignment
            # Write the value. 
            cell1 = ws.cell(row=i, column=2, value=value)
            cell1.fill = Res_fill
            cell1.border = cell_border
            cell1.font = cell_font
            cell1.alignment = left_alignment
        NextRowIndex = i + 2
        # ----------------------------------------------------------------------------------------------------------------------
        # Next, provide the deconvolution gaussians. 
        # Get the Gaussian lists. 
        ColNames = [ 
            'Deconv_GaussianList', 'Deconv_GaussianList_shape', 'Deconv_GaussianList_dtype', 
            'Deconv_CarbonylList', 'Deconv_CarbonylList_shape', 'Deconv_CarbonylList_dtype', 
            'Deconv_SulfoxideList', 'Deconv_SulfoxideList_shape', 'Deconv_SulfoxideList_dtype', 
            'Deconv_AliphaticList', 'Deconv_AliphaticList_shape', 'Deconv_AliphaticList_dtype']
        self.cursor.execute(f'SELECT {", ".join(ColNames)} FROM FTIR WHERE id = ?', (ID,))
        Content = list(self.cursor.fetchone())
        Gaussians = Binary_to_Array(Content[0], Content[1],  Content[2])
        GL_C      = Binary_to_Array(Content[3], Content[4],  Content[5])
        GL_S      = Binary_to_Array(Content[6], Content[7],  Content[8])
        GL_A      = Binary_to_Array(Content[9], Content[10], Content[11])
        # ------------------------------------------------------
        # Write the title. 
        ws.merge_cells(f'D1:H1')
        cell = ws.cell(row=1, column=4, value='Deconvolution Results')
        cell.fill = GL_fill
        cell.border = cell_border
        cell.font = Font(name="Arial", size=13, bold=True, color="000000")
        cell.alignment = center_alignment
        for i in range(2, 9):
            for j in range(5):
                cell1 = ws.cell(row=i, column=j+4, value='')
                cell1.fill = GL_fill
        # Put the image. 
        Image_Obj = Read_Resize_Image(ResourcePath(os.path.join(".", "assets", "Deconcolution Gaussian.png")), 125)
        ws.add_image(Image_Obj, "E2")
        NextRowIndex = 8
        Titles = ['Number', 'μ (cm⁻¹)', 'σ (cm⁻¹)', 'α', 'Area']
        for j in range(5):
            # Write the title.
            cell1 = ws.cell(row=NextRowIndex + 1, column=j + 4, value=Titles[j])
            cell1.fill = GL_fill
            cell1.border = cell_border
            cell1.font = header_font
            cell1.alignment = center_alignment
        for i in range(Gaussians.shape[0]):
            Values = [i + 1, Gaussians[i, 0], abs(Gaussians[i, 1]), Gaussians[i, 2], 
                    np.sqrt(2 * np.pi) * Gaussians[i, 2] * Gaussians[i, 1]]
            NumberFormat = ["0", "0.00", "0.0000", "0.000000", "0.0000"]
            for j in range(5):
                # Write the title.
                cell1 = ws.cell(row=i + NextRowIndex + 2, column=j + 4, value=Values[j])
                cell1.fill = GL_fill
                cell1.border = cell_border
                cell1.font = cell_font
                cell1.alignment = center_alignment
                cell1.number_format = NumberFormat[j]
        NextRowIndex = i + 2
        # ----------------------------------------------------------------------------------------------------------------------
        # Finally, Provide the raw data. 
        # Extract the raw data. 
        ColNames = [ 
            'Wavenumber', 'Wavenumber_shape', 'Wavenumber_dtype', 
            'Absorption', 'Absorption_shape', 'Absorption_dtype', 
            'RawWavenumber', 'RawWavenumber_shape', 'RawWavenumber_dtype', 
            'RawAbsorbance', 'RawAbsorbance_shape', 'RawAbsorbance_dtype',]
        self.cursor.execute(f'SELECT {", ".join(ColNames)} FROM FTIR WHERE id = ?', (ID,))
        Content = list(self.cursor.fetchone())
        X_BC = Binary_to_Array(Content[0], Content[1],  Content[2])     # Preprocessed: wavenumbers (cm⁻¹). 
        Y_BC = Binary_to_Array(Content[3], Content[4],  Content[5])     # Preprocessed: absorbance. 
        Xraw = Binary_to_Array(Content[6], Content[7],  Content[8])     # Raw data: wavenumbers (cm⁻¹). 
        Yraw = Binary_to_Array(Content[9], Content[10], Content[11])    # Raw data: absorbance.
        # ------------------------------------------------------
        # Write the title. 
        ws.merge_cells(f'J1:M1')
        cell = ws.cell(row=1, column=10, value='FTIR Test Data')
        cell.fill = Data_fill
        cell.border = cell_border
        cell.font = Font(name="Arial", size=13, bold=True, color="000000")
        cell.alignment = center_alignment
        ws.merge_cells(f'J2:K2')
        cell = ws.cell(row=2, column=10, value='Raw Data')
        cell.fill = Data_fill
        cell.border = cell_border
        cell.font = Font(name="Arial", size=13, bold=True, color="000000")
        cell.alignment = center_alignment
        ws.merge_cells(f'L2:M2')
        cell = ws.cell(row=2, column=12, value='Pre-Processed')
        cell.fill = Data_fill
        cell.border = cell_border
        cell.font = Font(name="Arial", size=13, bold=True, color="000000")
        cell.alignment = center_alignment
        Titles = ['Wavenumber (cm⁻¹)', 'Absorbance', 'Wavenumber (cm⁻¹)', 'Absorbance',]
        for j in range(4):
            # Write the title.
            cell1 = ws.cell(row=3, column=j + 10, value=Titles[j])
            cell1.fill = Data_fill
            cell1.border = cell_border
            cell1.font = header_font
            cell1.alignment = center_alignment
        for i in range(len(Xraw)):
            Values = [Xraw[i], Yraw[i], X_BC[i], Y_BC[i]]
            NumberFormat = ["0.00", "0.0000", "0.00", "0.0000"]
            for j in range(4):
                # Write the title.
                cell1 = ws.cell(row=i + 4, column=j + 10, value=Values[j])
                cell1.fill = Data_fill
                cell1.border = cell_border
                cell1.font = cell_font
                cell1.alignment = center_alignment
                cell1.number_format = NumberFormat[j]
        # --------------------------------------------------------------------------------------------------------------
        # Adjust the column dimensions. 
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 8
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 13
        ws.column_dimensions["G"].width = 13
        ws.column_dimensions["H"].width = 13
        ws.column_dimensions["I"].width = 8
        ws.column_dimensions["J"].width = 21
        ws.column_dimensions["K"].width = 15
        ws.column_dimensions["L"].width = 21
        ws.column_dimensions["M"].width = 15
        # Save the Excel file. 
        wb.save(os.path.join(Directory, FileName))
        # Return nothing. 
        return
    # ------------------------------------------------------------------------------------------------------------------
    def Check_Row_Selection(self, ActionLabel):
        """
        This function checks if a row from the table is selected and have valid data in it. Then, it will return the 
        row index and "id" of the selected row. 
        """
        # Find the selected index. 
        SelectedIndices = self.Table.selectionModel().selectedIndexes()
        if len(SelectedIndices) == 0:           
            # Nothing is selected. 
            QMessageBox.critical(self, "Data Selection Error!", 
                                 f"Row was not selected. Please first select the row you want to {ActionLabel} " + 
                                 f"from the database.")
            return -1, -1
        idx = SelectedIndices[0].row()
        # Check the id value. 
        ID = self.Table.item(idx, 0)
        if ID == None or ID.text() == '':
            # Table is empty. 
            QMessageBox.critical(self, "Data Selection Error!",
                                 f"Selected row ({idx + 1}) is empty. Please first fetch the data using the " +
                                 f'"Search and Filter" section, then select the intended row to {ActionLabel}, and ' + 
                                 f'then click the corresponding button.')
            return -1, -1
        else:
            # Return the row index and database "id" value correspond to the selected row. 
            return idx, int(ID.text())
# ======================================================================================================================
# ======================================================================================================================
# ======================================================================================================================


def Get_Color_4_COV(value):
    """
    This function gets the COV value and return a corresponding background color to specify high COV values. 

    :return: color number. 
    """
    # Ignore less than 15%.
    if value < 0.15:
        return "#f0f0f0"  # background color
    # Use "reds" colormap for between 15% to 50%.
    elif 0.15 <= value <= 0.5:
        # Normalize value to the range [0, 1] for the colormap
        normalized_value = (value - 0.15) / (0.5 - 0.15)
        # Get color from Reds colormap
        return to_hex(Custom_cmap(normalized_value))
    # Use "red" for more than 50%.
    else:
        return to_hex((1.0, 0.0, 0.0))  # Bright red
# ======================================================================================================================
# ======================================================================================================================
# ======================================================================================================================


def ResourcePath(relative_path):
    """
    Get the absolute path to the resource, works for dev and PyInstaller build.

    :param RelativePath: The relative path, which is going to be converted to the Resource Path. 
    """
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller stores resources in a temporary folder (_MEIPASS)
        return os.path.join(sys._MEIPASS, relative_path)
    else:
        # Use the relative path during development
        return os.path.join(os.path.abspath(relative_path))
# ======================================================================================================================
# ======================================================================================================================
# ======================================================================================================================


def Read_Resize_Image(path, targetPixel):
    """
    This function reads the image (*.png, *.jpg) and resize it to properly fit in the Excel file. 

    :param path: The complete/relative path to the image. 
    :param targetPixel: The height of the image after resize in pixels, given the fixed aspect ratio. 
    :param return: the resized image object. 
    """
    Image_Obj = Image(path)
    Ratio = targetPixel / Image_Obj.height
    Image_Obj.height = targetPixel
    Image_Obj.width  = Image_Obj.width * Ratio
    return Image_Obj
# ======================================================================================================================
# ======================================================================================================================
# ======================================================================================================================


if __name__ == '__main__':
    app = QApplication(sys.argv)
    # Connect to a SQL database.
    conn = sqlite3.connect(
        "C:\\Users\\SF.Abdollahi.ctr\\OneDrive - DOT OST\\FTIR_Temp\\PTF5_DB.db")
    cursor = conn.cursor()

    Main = DB_ReviewPage(conn, cursor, 'PTF5_DB')
    Main.show()
    app.exec()

    conn.close()
    app.quit()
    print('finish')
